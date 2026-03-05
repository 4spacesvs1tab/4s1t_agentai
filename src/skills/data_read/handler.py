#!/usr/bin/env python3
"""
data_read skill handler — reads and previews structured data files.

Supported formats: CSV, Excel (.xlsx/.xls), JSON (array or object), JSONL.

Sandboxed to FILE_READ_BASE_DIR (same path-traversal protection as file_read skill).

Input:  {"parameters": {"path": "...", "sample_rows": 10, "sheet_name": ""}}
Output: {"success": true, "result": {
    "columns": [...], "row_count": N,
    "sample": [...first N rows as dicts...],
    "dtypes": {"col": "type", ...},
    "file_format": "csv|excel|json|jsonl"
}}
"""
import csv
import json
import os
import sys
from pathlib import Path


# ---------------------------------------------------------------------------
# Sandbox helper (mirrors file_read skill)
# ---------------------------------------------------------------------------

def _resolve_path(requested: str) -> Path:
    base_dir = Path(os.environ.get("FILE_READ_BASE_DIR", "./data")).resolve()
    target = (base_dir / requested).resolve()
    if not str(target).startswith(str(base_dir)):
        raise PermissionError(
            f"Path traversal denied: '{requested}' resolves outside sandbox."
        )
    if not target.exists():
        raise FileNotFoundError(f"File not found in sandbox: '{requested}'")
    return target


# ---------------------------------------------------------------------------
# Format readers
# ---------------------------------------------------------------------------

def _infer_dtype(values: list) -> str:
    """Infer a simple dtype label from a sample of values (all already strings or Python objects)."""
    non_null = [v for v in values if v is not None and v != ""]
    if not non_null:
        return "null"
    # Try int
    try:
        for v in non_null[:20]:
            int(str(v))
        return "integer"
    except (ValueError, TypeError):
        pass
    # Try float
    try:
        for v in non_null[:20]:
            float(str(v))
        return "float"
    except (ValueError, TypeError):
        pass
    return "string"


def _read_csv(path: Path, sample_rows: int) -> dict:
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        columns = reader.fieldnames or []
        for row in reader:
            rows.append(dict(row))

    row_count = len(rows)
    sample = rows[:sample_rows]

    # Infer dtypes from full data
    dtypes = {}
    for col in columns:
        col_values = [r.get(col) for r in rows]
        dtypes[col] = _infer_dtype(col_values)

    return {
        "columns": list(columns),
        "row_count": row_count,
        "sample": sample,
        "dtypes": dtypes,
        "file_format": "csv",
    }


def _read_excel(path: Path, sample_rows: int, sheet_name: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required to read Excel files. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Sheet selection
    if sheet_name:
        if sheet_name.isdigit():
            idx = int(sheet_name)
            ws = wb.worksheets[idx]
        elif sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First row is header
    header_row = next(rows_iter, None)
    if header_row is None:
        return {
            "columns": [], "row_count": 0, "sample": [], "dtypes": {}, "file_format": "excel"
        }

    columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header_row)]

    all_rows = []
    for row in rows_iter:
        all_rows.append(dict(zip(columns, row)))

    wb.close()

    row_count = len(all_rows)
    sample = all_rows[:sample_rows]

    dtypes = {}
    for col in columns:
        col_values = [r.get(col) for r in all_rows]
        dtypes[col] = _infer_dtype(col_values)

    return {
        "columns": columns,
        "row_count": row_count,
        "sample": [
            {k: (v if v is None else str(v) if not isinstance(v, (int, float, bool)) else v)
             for k, v in row.items()}
            for row in sample
        ],
        "dtypes": dtypes,
        "file_format": "excel",
    }


def _read_json(path: Path, sample_rows: int) -> dict:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    # Normalise: accept array of objects or a single object
    if isinstance(data, dict):
        # Could be {"data": [...]} or a single record
        # Try to find an array key
        for v in data.values():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                data = v
                break
        else:
            data = [data]  # treat single object as one-row table

    if not data:
        return {"columns": [], "row_count": 0, "sample": [], "dtypes": {}, "file_format": "json"}

    if not isinstance(data[0], dict):
        # Array of scalars — wrap as {"value": ...}
        data = [{"value": v} for v in data]

    # Collect union of all keys
    columns = list(dict.fromkeys(k for row in data for k in row.keys()))
    row_count = len(data)
    sample = data[:sample_rows]

    dtypes = {}
    for col in columns:
        col_values = [r.get(col) for r in data]
        dtypes[col] = _infer_dtype(col_values)

    return {
        "columns": columns,
        "row_count": row_count,
        "sample": sample,
        "dtypes": dtypes,
        "file_format": "json",
    }


def _read_jsonl(path: Path, sample_rows: int) -> dict:
    all_rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            all_rows.append(json.loads(line))

    if not all_rows:
        return {"columns": [], "row_count": 0, "sample": [], "dtypes": {}, "file_format": "jsonl"}

    if not isinstance(all_rows[0], dict):
        all_rows = [{"value": v} for v in all_rows]

    columns = list(dict.fromkeys(k for row in all_rows for k in row.keys()))
    row_count = len(all_rows)
    sample = all_rows[:sample_rows]

    dtypes = {}
    for col in columns:
        col_values = [r.get(col) for r in all_rows]
        dtypes[col] = _infer_dtype(col_values)

    return {
        "columns": columns,
        "row_count": row_count,
        "sample": sample,
        "dtypes": dtypes,
        "file_format": "jsonl",
    }


# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------

_FORMAT_DISPATCH = {
    ".csv": _read_csv,
    ".xlsx": _read_excel,
    ".xls": _read_excel,
    ".json": _read_json,
    ".jsonl": _read_jsonl,
    ".ndjson": _read_jsonl,
}


def execute(params: dict) -> dict:
    requested = params.get("path", "")
    if not requested:
        raise ValueError("'path' parameter is required.")

    sample_rows = int(params.get("sample_rows", 10))
    sample_rows = max(1, min(sample_rows, 100))
    sheet_name = str(params.get("sheet_name", "") or "")

    target = _resolve_path(requested)
    suffix = target.suffix.lower()

    if suffix not in _FORMAT_DISPATCH:
        raise ValueError(
            f"Unsupported file format '{suffix}'. "
            f"Supported: {', '.join(_FORMAT_DISPATCH.keys())}"
        )

    reader = _FORMAT_DISPATCH[suffix]

    if suffix in (".xlsx", ".xls"):
        return reader(target, sample_rows, sheet_name)
    else:
        return reader(target, sample_rows)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) < 3:
        print("Usage: handler.py input.json output.json", file=sys.stderr)
        sys.exit(1)
    input_path, output_path = sys.argv[1], sys.argv[2]
    try:
        data = json.loads(open(input_path).read())
        params = data.get("parameters", {})
        result = execute(params)
        output = {"success": True, "result": result, "error": None, "logs": []}
    except Exception as exc:
        output = {"success": False, "result": None, "error": str(exc), "logs": []}
    with open(output_path, "w") as f:
        json.dump(output, f)


if __name__ == "__main__":
    main()
