#!/usr/bin/env python3
"""
export_document skill handler.

Assembles and exports a document in DOCX, PDF, PPTX, XLSX, HTML, or ODT format.

Format registry:
  docx / pdf / pptx / html / odt  →  _export_pandoc()   (requires: pandoc)
  xlsx                             →  _export_xlsx()     (requires: openpyxl pip package)

Files are saved to data/documents/.

Input:  {"parameters": {"title": "...", "format": "docx", "filename": "...", "sections": [...], "xlsx_sheets": null}}
Output: {"success": true, "result": {"file_id": "...", "download_path": "/api/v1/documents/...", ...}}
"""
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import uuid
from datetime import datetime
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
_DOCUMENTS_DIR = _PROJECT_ROOT / "data" / "documents"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sanitise_stem(name: str) -> str:
    """Turn a display name into a safe filename stem."""
    s = re.sub(r"[^\w\s\-]", "", name).strip()
    s = re.sub(r"[\s]+", "_", s)
    s = s[:60].strip("_") or "document"
    return s


def _unique_stem(stem: str) -> str:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    short_id = str(uuid.uuid4())[:6]
    return f"{stem}_{ts}_{short_id}"


_PANDOC_FMT = {
    "docx": "docx",
    "pdf":  "pdf",
    "pptx": "pptx",
    "html": "html5",
    "odt":  "odt",
}

_MIME = {
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "pdf":  "application/pdf",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "html": "text/html",
    "odt":  "application/vnd.oasis.opendocument.text",
}


# ---------------------------------------------------------------------------
# Pandoc handler (docx, pdf, pptx, html, odt)
# ---------------------------------------------------------------------------

def _build_markdown(title: str, sections: list) -> str:
    lines = [f"# {title}", ""]
    for sec in sections:
        heading = sec.get("heading", "")
        content = sec.get("content", "")
        if heading:
            lines.append(f"## {heading}")
            lines.append("")
        if content:
            lines.append(content)
            lines.append("")
    return "\n".join(lines)


def _export_pandoc(title: str, fmt: str, stem: str, sections: list, out_path: Path) -> None:
    pandoc = shutil.which("pandoc")
    if not pandoc:
        raise RuntimeError(
            "pandoc not found. Install with: apt-get install pandoc"
        )
    md_content = _build_markdown(title, sections or [])
    pandoc_fmt = _PANDOC_FMT[fmt]

    with tempfile.TemporaryDirectory() as tmpdir:
        md_file = Path(tmpdir) / "input.md"
        md_file.write_text(md_content, encoding="utf-8")

        cmd = [pandoc, str(md_file), "-o", str(out_path), "--to", pandoc_fmt]
        if fmt == "html":
            cmd += ["--standalone", "--embed-resources"]
        if fmt == "pdf":
            # Prefer wkhtmltopdf if available (no LaTeX needed), else let pandoc pick
            if shutil.which("wkhtmltopdf"):
                cmd += ["--pdf-engine", "wkhtmltopdf"]
            elif shutil.which("weasyprint"):
                cmd += ["--pdf-engine", "weasyprint"]

        # Run from /tmp — /app/src is a read-only volume mount; pandoc needs a writable cwd
        # for temp files it creates during conversion.
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=55,
                                cwd=tempfile.gettempdir())
        if result.returncode != 0:
            raise RuntimeError(f"pandoc failed: {result.stderr.strip() or result.stdout.strip()}")


# ---------------------------------------------------------------------------
# XLSX handler (openpyxl)
# ---------------------------------------------------------------------------

def _export_xlsx(title: str, stem: str, xlsx_sheets: list, out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError(
            "openpyxl not installed. Install with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D5A8E", end_color="2D5A8E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for sheet_def in xlsx_sheets:
        ws = wb.create_sheet(title=sheet_def.get("name", "Sheet"))
        headers = sheet_def.get("headers", [])
        rows = sheet_def.get("rows", [])
        col_widths = sheet_def.get("col_widths") or []

        # Write header row
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Set column widths
        for col_idx, width in enumerate(col_widths, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # Auto-width for columns without explicit width
        for col_idx in range(len(col_widths) + 1, len(headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1]) if col_idx <= len(headers) else 10
            for row in rows:
                if col_idx - 1 < len(row):
                    val_len = len(str(row[col_idx - 1] or ""))
                    max_len = max(max_len, val_len)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

    wb.save(str(out_path))


# ---------------------------------------------------------------------------
# Format registry
# ---------------------------------------------------------------------------

_PANDOC_FORMATS = {"docx", "pdf", "pptx", "html", "odt"}


def execute(params: dict) -> dict:
    title = (params.get("title") or "Document").strip()
    fmt = (params.get("format") or "docx").strip().lower()
    if fmt not in _MIME:
        raise ValueError(f"Unsupported format '{fmt}'. Supported: {', '.join(_MIME)}")

    stem = _sanitise_stem(params.get("filename") or title)
    unique_stem = _unique_stem(stem)
    file_id = f"{unique_stem}.{fmt}"

    _DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
    out_path = _DOCUMENTS_DIR / file_id

    if fmt in _PANDOC_FORMATS:
        sections = params.get("sections") or []
        _export_pandoc(title, fmt, unique_stem, sections, out_path)
    elif fmt == "xlsx":
        xlsx_sheets = params.get("xlsx_sheets")
        if not xlsx_sheets:
            raise ValueError("'xlsx_sheets' is required for xlsx format")
        _export_xlsx(title, unique_stem, xlsx_sheets, out_path)
    else:
        raise ValueError(f"No handler registered for format '{fmt}'")

    if not out_path.exists():
        raise RuntimeError(f"Export produced no output file at {out_path}")

    size = out_path.stat().st_size
    return {
        "file_id": file_id,
        "download_path": f"/api/v1/documents/{file_id}",
        "format": fmt,
        "filename": f"{stem}.{fmt}",
        "size_bytes": size,
    }


# ---------------------------------------------------------------------------
# Subprocess entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) < 3:
        print("Usage: handler.py input.json output.json", file=sys.stderr)
        sys.exit(1)
    input_path, output_path = sys.argv[1], sys.argv[2]
    try:
        data = json.loads(open(input_path).read())
        params = data.get("parameters", {})
        result = execute(params)
        output = {"success": True, "result": result, "error": None, "logs": []}
    except Exception as exc:
        output = {"success": False, "result": None, "error": str(exc), "logs": []}
    with open(output_path, "w") as f:
        json.dump(output, f)


if __name__ == "__main__":
    main()
